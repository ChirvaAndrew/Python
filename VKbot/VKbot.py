import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
import requests
from bs4 import BeautifulSoup
import openpyxl
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import datetime
import json


def sevendays(s): #Возвращает дату начала и конца текущей недели в формате День.Месяц.Год - День.Месяц.Год
    current_date = datetime.datetime.now()
    day = current_date.strftime('%m.%d.%y')
    dt = datetime.datetime.strptime(day, '%m.%d.%y')
    start = dt - datetime.timedelta(days = dt.weekday()) - datetime.timedelta(days=s)#Начало недели
    return start
#Возвращает текущее время
def nowp(a,n):#Аргументы: 0 - Дата/Месяц/Год; 1 - Дата; 2 - Месяц; 3 - Год
    current_date = datetime.datetime.now()
    if a == 0:
        return current_date.strftime('%d.%m')
    if a == 1:
        return current_date.strftime('%d')
    if a == 2:
        current_date = datetime.datetime.today() + datetime.timedelta(days=n)
        return current_date.strftime('%d.%m')
    if a == 3:
        return current_date.strftime('%Y')
    if a == 4:
        return current_date.weekday()


def write_msg(user_id, message):
    random_id = vk_api.utils.get_random_id()
    vk.method('messages.send', {'user_id': user_id, 'message': message, "random_id":random_id})


# API-ключ созданный ранее
token = "090cd1aa05d0c1db4c38d8a8f74f3d38ab85c82ab1c570ffebdb94562b0e19a6dddafd983adfb02bd6470"

# Авторизуемся как сообщество
vk = vk_api.VkApi(token=token)
vkplus = vk.get_api()
# Работа с сообщениями
longpoll = VkLongPoll(vk)
week = datetime.datetime.today() - datetime.datetime(2022, 2, 9)
week = int (week.days / 7 + 1)

def Rasp(group):
    result = ""
    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")
    for link in soup.find("div", {"class":"rasspisanie"}).\
        find(string = "Институт информационных технологий").\
        find_parent("div").\
        find_parent("div").\
        find_all('a', href=True):
            result += "\n" + str(link['href'])
    #print(result)
    my_file = open("Links.txt", "w+")
    my_file.write(result)
    my_file.close()
    course = str(22 - int(group[-2::]))
    l = open("Links.txt")
    l.readline()
    l.readline()
    while(True):
        line = l.readline()
        if not line:
            break
        if line[84] == course:
            linkf = line
            break
    l.close()
    linkf = linkf[:-1]
    #print(linkf)
    f = open("file.xlsx", "wb")  # открываем файл для записи, в режиме wb
    resp = requests.get(linkf)  # запрос по ссылке
    f.write(resp.content)


def finder():
    book = openpyxl.load_workbook("file.xlsx") # открытие файла
    sheet = book.active # активный лист
    num_cols = sheet.max_column # количество столбцов
    num_rows = sheet.max_row # количество строк
    cell = sheet.cell(row = 2, column = 16).value # ячейка
    #print(cell)
def sbj(i,col, book):
    sheet = book.active
    return str(sheet.cell(row=i, column=col).value) + "," + str(sheet.cell(row=i, column= col + 1).value + "," + str(sheet.cell(row=i, column= col + 2).value)+ "," + str(sheet.cell(row=i, column= col + 3).value))



def keboard(group):
    group = group.upper()
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('На завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()  # переход на вторую строку
    keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()  # переход на вторую строку
    keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
    keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)

    vkplus.messages.send(user_id=event.user_id, random_id=vk_api.utils.get_random_id(),keyboard=keyboard.get_keyboard(), message='показать расписание ' + group.upper())

def today(group,tomorrow,ch):
    Rasp(group)
    ans =""
    global week
    tempe = week
    if (ch != 0 and tomorrow == 1):
        tempe = week
        week += 1
    group = group.upper()
    book = openpyxl.load_workbook("file.xlsx")  # открытие файла
    sheet = book.active  # активный лист
    num_cols = sheet.max_column  # количество столбцов
    num_rows = sheet.max_row  # количество строк
    #cell = sheet.cell(row=16, column=16).value  # ячейка
    #print(type(cell))
    for i in range(1,num_cols):
        if sheet.cell(row=2, column=i).value == group:
            col = i
            break
    if ch == 0:
        day = nowp(4,0) + 1 + tomorrow
    else:
        day = ch
    if day == 1:
        c = 1
        if week % 2 == 1:
            for i in range(4, 15, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(5, 16, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    if day == 2:
        c = 1
        if week % 2 == 1:
            for i in range(16, 27, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(17, 28, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    if day == 3:
        c = 1
        if week % 2 == 1:
            for i in range(28, 39, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(29, 40, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    if day == 4:
        c = 1
        if week % 2 == 1:
            for i in range(40, 51, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(41, 52, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    if day == 5:
        c = 1
        if week % 2 == 0:
            for i in range(53, 64, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(52, 63, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    if day == 6:
        c = 1
        if week % 2 == 1:
            for i in range(64, 75, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c)+")"+sub+'\n'
                c += 1
        else:
            for i in range(65, 76, 2):
                sub = ""
                if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                    sub = "—"
                else:
                    sub = sbj(i,col, book)
                ans += str(c) + ")" + sub+"\n"
                c += 1
    ans = ans.replace("None", "—")
    if (ch == 0):
        if (tomorrow == 0):
            write_msg(event.user_id, "Расписание на " + nowp(0,0) + ": \n" + ans)
        if (tomorrow == 1):
            write_msg(event.user_id, "Расписание на " + nowp(2,1) + ": \n" + ans)
    else:
        if(week % 2 == 0):
            write_msg(event.user_id, "Расписание на чётную неделю: " + "\n" + ans)
        else:
            write_msg(event.user_id, "Расписание на нечётную неделю: " + "\n" + ans)
    week = tempe

def AllW(group,tomorrow):
    Rasp(group)
    ans = ""
    ansb = ""
    global week
    group = group.upper()
    book = openpyxl.load_workbook("file.xlsx")  # открытие файла
    sheet = book.active  # активный лист
    num_cols = sheet.max_column  # количество столбцов
    num_rows = sheet.max_row  # количество строк
    # cell = sheet.cell(row=16, column=16).value  # ячейка
    # print(type(cell))
    for i in range(1, num_cols):
        if sheet.cell(row=2, column=i).value == group:
            col = i
            break
    tempo = week
    if tomorrow == 1:
        week += 1
    for day in range (1,7):
        ans = ""
        if day == 1:
            c = 1
            if week % 2 == 1:
                for i in range(4, 15, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(5, 16, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or sheet.cell(row=i, column=col).value== "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        if day == 2:
            c = 1
            if week % 2 == 1:
                for i in range(16, 27, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(17, 28, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        if day == 3:
            c = 1
            if week % 2 == 1:
                for i in range(28, 39, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(29, 40, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        if day == 4:
            c = 1
            if week % 2 == 1:
                for i in range(40, 51, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(41, 52, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        if day == 5:
            c = 1
            if week % 2 == 0:
                for i in range(53, 64, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(52, 63, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        if day == 6:
            c = 1
            if week % 2 == 1:
                for i in range(64, 75, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + '\n'
                    c += 1
            else:
                for i in range(65, 76, 2):
                    sub = ""
                    if sheet.cell(row=i, column=col).value is None or str(sheet.cell(row=i, column=col).value) == "…………………..":
                        sub = "—"
                    else:
                        sub = sbj(i, col, book)
                    ans += str(c) + ")" + sub + "\n"
                    c += 1
        ans = ans.replace("None", "—")
        if tomorrow == 0:
            x = sevendays(1) + datetime.timedelta(days=day)
            ansb += "Расписание на " + x.strftime('%d.%m') + "\n" + ans +"\n\n"
        else:
            x = sevendays(-6) + datetime.timedelta(days=day)
            ansb += "Расписание на " + x.strftime('%d.%m') + "\n" + ans + "\n\n"
    write_msg(event.user_id, ansb)
    week = tempo




def moscow():
    translate = {'Thunderstorm': "Гроза", 'Drizzle': "Моросит", 'Rain': "Дождь", 'Snow': "Снег", 'Mist': "Туман", 'Smoke': "Дым", 'Haze': "Легкий туман", 'Dust': "Пыль", 'Fog': "Туман", 'Sand': "Моросит", 'Песок': "Clear", 'Ясно': "Моросит", 'Clouds': "Облачно", 'Tornado': "торнадо", 'Squall': "Шквал", 'Ash': "Пепел",
                 'clear sky':"чистое небо", 'few clouds':"несколько облаков",'scattered clouds':"рассеянные облака", 'overcast clouds':"ААА",'broken clouds':"разорванные облака",'shower rain':"ливень",'rain':"дождь",'thunderstorm':"гроза",'snow':"снег",'mist':"туман",'thunderstorm with light rain':"гроза с небольшим дождем",'thunderstorm with rain':"гроза с дождем",'thunderstorm with heavy rain':"гроза с ливнем",'light thunderstorm':"небольшая гроза",'heavy thunderstorm':"сильная гроза",'ragged thunderstorm':"ураган",
                 'thunderstorm with light drizzle':"гроза с небольшим туманом",'thunderstorm with drizzle':"гроза с туманом",'thunderstorm with heavy drizzle':"гроза с сильным туманом",'light intensity drizzle':"незначительный туман",'heavy intensity drizzle':"интенсивный туман",'light intensity drizzle rain':"легкий дождь с туманом",'drizzle rain':"дождь с туманом",'heavy intensity drizzle rain':"сильный туман с дождём",'shower rain and drizzle':"туман и ливень",'heavy shower rain and drizzle':"сильный туман и ливень",'shower drizzle':"туман с ураганом",'light rain':"небольшой дождь",'moderate rain':"умеренный дождь",'heavy intensity rain':"сильный дождь",'very heavy rain':"очень сильный дождь",'extreme rain':"черезвычайно сильный дождь",
                 'freezing rain':"ледяной дождь",'light intensity shower rain':"небольшой ледяной дождь",'heavy intensity shower rain':"сильный ливень",'ragged shower rain':"ледянной ливень",'light snow':"легкий снег",'Heavy snow':"сильный снег",'Sleet':"Слякоть",'Light shower sleet':"Легкий дождь с мокрым снегом",'Shower sleet':"Ливень с мокрым снегом",'Light rain and snow':"легкий дождь со снегом",'Rain and snow':"дождь со снегом",'Light shower snow':"Легкий снежный дождь",'Shower snow':"Ливень снега",'Heavy shower snow':"Сильный ливень со снегом",'sand/ dust whirls':"вихри песка/ пыли",'volcanic ash':"Вулканический пепел",
                 'few clouds: 11-25%':"мало облаков: 11-25%",'scattered clouds: 25-50%':"рассеянные облака: 25-50%",'broken clouds: 51-84%':"разорванные облака: 51-84%",'overcast clouds: 85-100%':"Облачность: 85-100%"}
    weather = requests.get("http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=bc6fb75f3340b3fbb4417fd96406e0f1&units=metric")
    jweather = json.loads(weather.content)
    for i in jweather["weather"]:
        mainW = (i["main"])
        mainW = (translate[mainW])
        descriptionW = (i["description"])
        descriptionW = (translate[descriptionW])
    temp_min = jweather["main"]["temp_min"]
    temp_max = jweather["main"]["temp_max"]
    pressure = jweather["main"]["pressure"]
    aqua = jweather["main"]["humidity"]
    speed = jweather["wind"]["speed"]
    deg = jweather["wind"]["deg"]
    winddirections = ("северный", "северо-восточный", "восточный", "юго-восточный", "южный", "юго-западный", "западный",
                      "северо-западный")
    direction = int((deg + 22.5) // 45 % 8)
    dir = str(winddirections[direction])
    if float(speed) <= 0.2:
        windtype = "Штиль"
    if float(speed) >= 0.3 and float(speed) <= 1.5:
        windtype = "Тихий"
    if float(speed) >= 1.6 and float(speed) <= 3.3:
         windtype = "Лёгкий"
    if float(speed) >= 3.4 and float(speed) <= 5.4:
        windtype = "Слабый"
    if float(speed) >= 5.5 and float(speed) <= 7.9:
        windtype = "Умеренный"
    if float(speed) >= 8 and float(speed) <= 10.7:
        windtype = "Свежий"
    if float(speed) >= 10.8 and float(speed) <= 13.8:
        windtype = "Сильный"
    if float(speed) >= 13.9 and float(speed) <= 17.1:
        windtype = "Крепкий"
    if float(speed) >= 17.2 and float(speed) <= 20.7:
        windtype = "Очень крепкий"
    if float(speed) >= 20.8 and float(speed) <= 24.4:
        windtype = "Шторм"
    if float(speed) >= 24.5 and float(speed) <= 28.4:
        windtype = "Сильный шторм"
    if float(speed) >= 28.5 and float(speed) <= 32.6:
        windtype = "Жестокий шторм"
    if float(speed) >= 33:
        windtype = "Ураган"
    forecat = ("Погода в Москве: " + mainW + "\n" + descriptionW + " Температура " + str(int(temp_min)) + "-" + str(int(temp_max)) + "°C\n" +"Давление: " + str(int(float(pressure) * 0.750064)) + "мм рт.ст. Влажность: " + str(aqua) + "%\n" +"Ветер: " + windtype +", "+ str(speed) + "м/с, " + dir)
    """
    print("Погода в Москве: " + mainW)
    print(descriptionW + " Температура " + str(int(temp_min)) + "-" + str(int(temp_max)) + "°C")
    print("Давление: " + str(int(float(pressure) * 0.750064)) + "мм рт.ст. Влажность: " + str(aqua) + "%")
    print("Ветер: " + windtype +)
    """
    write_msg(event.user_id, forecat)





















group = "ИКБО-31-21"
finder()
# Основной цикл
for event in longpoll.listen():

    # Если пришло новое сообщение
    if event.type == VkEventType.MESSAGE_NEW:

        # Если оно имеет метку для меня(то есть бота)
        if event.to_me:
            print('New from {}, text = {}'.format(event.user_id, event.text))
            # Сообщение от пользователя
            request = event.text
            if len(request) == 10 and request[4] == "-" and request[7] == "-":
                write_msg(event.user_id, "Я запомнил, что вы из " + request.upper())
                group = request
                Rasp(group)
                continue
            request = request.lower()
            # Каменная логика ответа
            if request == "начать":
                write_msg(event.user_id, "Здравствуйте, вы можете указать свою группу(По умолчанию ИКБО-31-21) или посмотреть расписание с помощью команды 'бот'\n Также я могу показать погоду в Москве по команде'погода'")
                continue
            if request == "привет":
                write_msg(event.user_id, 'Привет, ' + vkplus.users.get(user_id = event.user_id)[0]['first_name'])
                continue
            elif request == "пока":
                write_msg(event.user_id, "Пока((")
                continue
            elif request == "бот":
                keboard(group)
                continue
            elif request[:4:] == "бот " and request[8] == "-" and request[11] == "-":
                group = request[4::]
                keboard(group)
                continue
            elif request == "на сегодня":
                today(group,0,0)
                continue
            elif request == "на завтра":
                today(group,1,0)
                continue
            elif request == "на эту неделю":
                AllW(group, 0)
                continue
            elif request == "погода":
                moscow()
                continue
            elif request == "на следующую неделю":
                AllW(group, 1)
                continue
            elif request == "какая группа?":
                write_msg(event.user_id, "Показываю расписание группы " + group.upper())
                continue
            elif request == "какая неделя?":
                write_msg(event.user_id, "Идёт " + str(week) + " неделя")
                continue


            if request[:4:] == "бот " and request[8] != "-" and request[len(request) - 3] == "-":
                group = request[-10::]
                request = request[:-11:]
                request = request[4::]
                if request == "понедельник":
                    today(group, 0, 1)
                    today(group, 1, 1)
                if request == "вторник":
                    today(group, 0, 2)
                    today(group, 1, 2)
                if request == "среда":
                    today(group, 0, 3)
                    today(group, 1, 3)
                if request == "четверг":
                    today(group, 0, 4)
                    today(group, 1, 4)
                if request == "пятница":
                    today(group, 0, 5)
                    today(group, 1, 5)
                if request == "суббота":
                    today(group, 0, 6)
                    today(group, 1, 6)
                continue

            if request[:4:] == "бот " and request[8]!= "-":
                request = request[4::]
                if request == "понедельник":
                    today(group,0,1)
                    today(group, 1, 1)
                if request == "вторник":
                    today(group,0,2)
                    today(group, 1, 2)
                if request == "среда":
                    today(group,0,3)
                    today(group, 1, 3)
                if request == "четверг":
                    today(group,0,4)
                    today(group, 1, 4)
                if request == "пятница":
                    today(group,0,5)
                    today(group, 1, 5)
                if request == "суббота":
                    today(group,0,6)
                    today(group, 1, 6)
                continue
            else:
                write_msg(event.user_id, "Неизвестная команда")


